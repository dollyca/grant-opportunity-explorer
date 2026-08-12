import streamlit as st
import pandas as pd
from datetime import date
from pathlib import Path
from PIL import Image
from openpyxl import load_workbook


# --------------------------------------------------
# Page setup
# --------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "data" / "Grant Database - Backup.xlsx"
LOGO_FILE = BASE_DIR / "medica_logo.PNG"

logo = Image.open(LOGO_FILE)

st.set_page_config(
    page_title="Grant Opportunity Explorer",
    page_icon=logo,
    layout="wide",
)


# --------------------------------------------------
# Helpers
# --------------------------------------------------

CORE_COLUMNS = [
    "Funder Name",
    "Program Name",
    "Website Link",
    "Funding Type",
    "Focus Area",
    "Grant Size",
    "Eligibility",
    "Deadline",
    "Deadline Type",
    "Geographic Scope",
    "Notes (Fit with Medica Zone's Mission)",
]


def clean_text(series):
    """Trim extra whitespace while preserving missing values."""
    return (
        series.astype("string")
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
        .replace("", pd.NA)
    )


def extract_excel_hyperlinks(file_path, sheet_name):
    """
    Read the actual hyperlink targets stored in the Excel Website Link column.
    Pandas normally reads only the visible cell text.
    """
    workbook = load_workbook(file_path, data_only=False)
    worksheet = workbook[sheet_name]

    headers = {
        str(cell.value).strip(): index
        for index, cell in enumerate(worksheet[1], start=1)
        if cell.value is not None
    }

    website_col = headers.get("Website Link")

    if website_col is None:
        return []

    links = []

    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=website_col)

        if cell.hyperlink and cell.hyperlink.target:
            links.append(cell.hyperlink.target)
        elif isinstance(cell.value, str) and cell.value.startswith(
            ("http://", "https://")
        ):
            links.append(cell.value.strip())
        else:
            links.append(pd.NA)

    return links


def prepare_deadlines(df):
    """
    Deadline contains actual calendar dates only.
    Deadline Type stores the timing rule:
    Fixed, Rolling, Multiple, Annual, Estimated, or Varies.
    """
    if "Deadline" not in df.columns:
        df["Deadline"] = pd.NaT

    if "Deadline Type" not in df.columns:
        df["Deadline Type"] = pd.NA

    df["Deadline"] = pd.to_datetime(
        df["Deadline"],
        errors="coerce",
    )

    df["Deadline Type"] = clean_text(df["Deadline Type"])

    df["Deadline Display"] = (
        df["Deadline"]
        .dt.strftime("%m/%d/%Y")
        .fillna("")
    )

    today = pd.Timestamp(date.today())

    df["Days Until Deadline"] = (
        df["Deadline"].dt.normalize() - today
    ).dt.days

    def classify_deadline(row):
        deadline = row["Deadline"]
        deadline_type = row["Deadline Type"]
        days = row["Days Until Deadline"]

        if pd.isna(deadline):
            if pd.notna(deadline_type):
                return str(deadline_type)
            return "Missing deadline"

        if days < 0:
            return "Past deadline"

        if days <= 7:
            return "Due within 7 days"

        if days <= 30:
            return "Due within 30 days"

        if days <= 90:
            return "Due within 90 days"

        return "More than 90 days"

    df["Deadline Status"] = df.apply(
        classify_deadline,
        axis=1,
    )

    return df


def add_missing_information_flag(df):
    """
    Flag rows that have at least one blank field in the core source data.
    Deadline is treated as complete when Deadline Type explains why
    no exact date exists (for example Rolling or Varies).
    """
    columns_to_check = [
        column
        for column in CORE_COLUMNS
        if column in df.columns
    ]

    if not columns_to_check:
        df["Has Missing Information"] = False
        return df

    missing_flags = []

    for _, row in df.iterrows():
        row_has_missing = False

        for column in columns_to_check:
            value = row[column]

            # A blank Deadline is acceptable when a Deadline Type exists.
            if column == "Deadline":
                if pd.isna(value) and pd.notna(row.get("Deadline Type")):
                    continue

            # Deadline Type itself can be blank when a fixed date exists.
            if column == "Deadline Type":
                if pd.isna(value) and pd.notna(row.get("Deadline")):
                    continue

            if pd.isna(value):
                row_has_missing = True
                break

            if isinstance(value, str) and not value.strip():
                row_has_missing = True
                break

        missing_flags.append(row_has_missing)

    df["Has Missing Information"] = missing_flags

    return df


# --------------------------------------------------
# Load data
# --------------------------------------------------

@st.cache_data
def load_data():
    # Read the first worksheet.
    excel_file = pd.ExcelFile(DATA_FILE)
    sheet_name = excel_file.sheet_names[0]

    df = pd.read_excel(
        DATA_FILE,
        sheet_name=sheet_name,
    )

    # Standardize header spacing only.
    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
    )

    # Remove completely empty rows.
    df = df.dropna(
        how="all"
    ).reset_index(drop=True)

    # Lightweight whitespace cleanup only.
    for column in df.columns:
        if column != "Deadline":
            df[column] = (
                clean_text(df[column])
                if df[column].dtype == "object"
                or str(df[column].dtype).startswith("string")
                else df[column]
            )

    # Replace visible Excel hyperlink labels with real URLs.
    website_links = extract_excel_hyperlinks(
        DATA_FILE,
        sheet_name,
    )

    if "Website Link" in df.columns and website_links:
        df["Website Link"] = pd.Series(
            website_links[: len(df)],
            dtype="string",
        )

    df = prepare_deadlines(df)
    df = add_missing_information_flag(df)

    return df


df = load_data()


# --------------------------------------------------
# Header
# --------------------------------------------------

st.title("Grant Opportunity Explorer")

st.caption(
    "Explore, filter, and analyze grant opportunities researched "
    "by the Budgeting & Finance team."
)


# --------------------------------------------------
# KPI cards
# --------------------------------------------------

total_grants = len(df)

funding_type_count = (
    df["Funding Type"].nunique(dropna=True)
    if "Funding Type" in df.columns
    else 0
)

focus_area_count = (
    df["Focus Area"].nunique(dropna=True)
    if "Focus Area" in df.columns
    else 0
)

missing_count = int(
    df["Has Missing Information"].sum()
)

k1, k2, k3, k4 = st.columns(4)

k1.metric(
    "Total Grants",
    total_grants,
)

k2.metric(
    "Funding Types",
    funding_type_count,
)

k3.metric(
    "Focus Areas",
    focus_area_count,
)

k4.metric(
    "Records with Missing Fields",
    missing_count,
    help=(
        "Number of grant records with at least one missing core field. "
        "A blank deadline is not counted as missing when a valid "
        "Deadline Type such as Rolling or Varies is provided."
    ),
)

st.divider()


# --------------------------------------------------
# Filters
# --------------------------------------------------

st.subheader("Explore Grants")

funding_options = (
    sorted(
        df["Funding Type"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
    if "Funding Type" in df.columns
    else []
)

focus_options = (
    sorted(
        df["Focus Area"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
    if "Focus Area" in df.columns
    else []
)

deadline_options = (
    sorted(
        df["Deadline Status"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )
    if "Deadline Status" in df.columns
    else []
)

f1, f2, f3 = st.columns(3)

selected_funding = f1.multiselect(
    "Funding Type",
    options=funding_options,
)

selected_focus = f2.multiselect(
    "Focus Area",
    options=focus_options,
)

selected_deadline_status = f3.multiselect(
    "Deadline Status",
    options=deadline_options,
)

show_missing = st.checkbox(
    "Show only grants with missing information"
)

search_term = st.text_input(
    "Search grants",
    placeholder=(
        "Search funder, program, focus area, eligibility, "
        "geographic scope, grant size, or notes..."
    ),
)


# --------------------------------------------------
# Apply filters
# --------------------------------------------------

filtered_df = df.copy()

if selected_funding and "Funding Type" in filtered_df.columns:
    filtered_df = filtered_df[
        filtered_df["Funding Type"].isin(
            selected_funding
        )
    ]

if selected_focus and "Focus Area" in filtered_df.columns:
    filtered_df = filtered_df[
        filtered_df["Focus Area"].isin(
            selected_focus
        )
    ]

if selected_deadline_status:
    filtered_df = filtered_df[
        filtered_df["Deadline Status"].isin(
            selected_deadline_status
        )
    ]

if show_missing:
    filtered_df = filtered_df[
        filtered_df["Has Missing Information"]
    ]

if search_term:
    searchable_columns = [
        "Funder Name",
        "Program Name",
        "Funding Type",
        "Focus Area",
        "Grant Size",
        "Eligibility",
        "Deadline Type",
        "Geographic Scope",
        "Notes (Fit with Medica Zone's Mission)",
    ]

    searchable_columns = [
        column
        for column in searchable_columns
        if column in filtered_df.columns
    ]

    if searchable_columns:
        search_mask = (
            filtered_df[searchable_columns]
            .fillna("")
            .astype(str)
            .apply(
                lambda column: column.str.contains(
                    search_term,
                    case=False,
                    na=False,
                    regex=False,
                )
            )
            .any(axis=1)
        )

        filtered_df = filtered_df[
            search_mask
        ]

st.write(
    f"Showing **{len(filtered_df)}** of "
    f"**{total_grants}** grants"
)


# --------------------------------------------------
# Grant table
# --------------------------------------------------

table_df = filtered_df.copy()

columns_to_show = [
    "Funder Name",
    "Program Name",
    "Website Link",
    "Funding Type",
    "Focus Area",
    "Grant Size",
    "Eligibility",
    "Deadline Display",
    "Deadline Type",
    "Days Until Deadline",
    "Deadline Status",
    "Geographic Scope",
    "Notes (Fit with Medica Zone's Mission)",
]

columns_to_show = [
    column
    for column in columns_to_show
    if column in table_df.columns
]

table_df = table_df[
    columns_to_show
].copy()

table_df = table_df.rename(
    columns={
        "Deadline Display": "Deadline",
        "Days Until Deadline": "Days Left",
        "Notes (Fit with Medica Zone's Mission)": "Mission Fit Notes",
    }
)

text_columns = [
    column
    for column in table_df.columns
    if column != "Days Left"
]

table_df[text_columns] = (
    table_df[text_columns]
    .fillna("")
    .astype(str)
)

st.dataframe(
    table_df,
    width="stretch",
    hide_index=True,
    column_config={
        "Website Link": st.column_config.LinkColumn(
            "Website Link",
            display_text="Open grant page",
        ),
        "Days Left": st.column_config.NumberColumn(
            "Days Left",
            format="%d",
        ),
        "Mission Fit Notes": st.column_config.TextColumn(
            "Mission Fit Notes",
            width="large",
        ),
    },
)


# --------------------------------------------------
# Download filtered results
# --------------------------------------------------

csv_data = table_df.to_csv(
    index=False
).encode("utf-8-sig")

st.download_button(
    label="Download Filtered Results as CSV",
    data=csv_data,
    file_name="filtered_grant_opportunities.csv",
    mime="text/csv",
)

st.divider()


# --------------------------------------------------
# Charts
# --------------------------------------------------

st.subheader("Grant Portfolio Overview")

# Keep the chart narrower so it is easier to read.
chart_col, empty_col = st.columns([3, 2])

with chart_col:
    st.markdown("#### Grants by Funding Type")

    funding_chart = (
        filtered_df["Funding Type"]
        .fillna("Missing")
        .astype(str)
        .value_counts()
        .rename_axis("Funding Type")
        .reset_index(name="Number of Grants")
    )

    if funding_chart.empty:
        st.info("No data available for the current filters.")
    else:
        st.bar_chart(
            funding_chart,
            x="Funding Type",
            y="Number of Grants",
            width="stretch",
        )